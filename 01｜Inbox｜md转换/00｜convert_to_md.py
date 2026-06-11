from __future__ import annotations

import csv
import re
import sys
from datetime import date, datetime
from pathlib import Path
from xml.sax.saxutils import escape

from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader


ROOT = Path(__file__).resolve().parent
OUT_DIR = ROOT / "已转换"
SKIP_DIRS = {"已转换", "原文件归档"}
SUPPORTED = {".docx", ".xlsx", ".xlsm", ".csv", ".tsv", ".txt", ".rtf", ".pdf"}


def clean_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    text = text.replace("|", "\\|")
    text = "<br>".join(line.strip() for line in text.split("\n"))
    return text.strip()


def markdown_table(rows: list[list[str]]) -> str:
    rows = [row for row in rows if any(cell.strip() for cell in row)]
    if not rows:
        return "_空表_"
    width = max(len(row) for row in rows)
    rows = [row + [""] * (width - len(row)) for row in rows]
    header = rows[0]
    if all(not cell for cell in header):
        header = [f"列{i + 1}" for i in range(width)]
        body = rows[1:]
    else:
        header = [cell if cell else f"列{i + 1}" for i, cell in enumerate(header)]
        body = rows[1:]
    lines = [
        "| " + " | ".join(header) + " |",
        "|" + "|".join("---" for _ in header) + "|",
    ]
    for row in body:
        lines.append("| " + " | ".join(row) + " |")
    return "\n".join(lines)


def frontmatter(src: Path) -> list[str]:
    return [
        f"# {src.stem}",
        "",
        "## 一、来源",
        "",
        f"源文件：`{src.name}`",
        "",
        "转换说明：由 Inbox 转 Markdown 工具生成。原始文件未移动、未删除。",
        "",
        "------------------------------------------------------------------------",
        "",
    ]


def convert_docx(src: Path) -> str:
    doc = Document(src)
    parts = frontmatter(src)
    parts.extend(["## 二、正文", ""])
    table_index = 1
    for block in doc.element.body:
        tag = block.tag.split("}")[-1]
        if tag == "p":
            text = "".join(node.text or "" for node in block.iter() if node.tag.split("}")[-1] == "t").strip()
            if text:
                parts.extend([text, ""])
        elif tag == "tbl":
            table = doc.tables[table_index - 1]
            rows = []
            for row in table.rows:
                rows.append([clean_cell(cell.text) for cell in row.cells])
            parts.extend([f"### 表格 {table_index}", "", markdown_table(rows), ""])
            table_index += 1
    return "\n".join(parts).rstrip() + "\n"


def convert_xlsx(src: Path) -> str:
    wb = load_workbook(src, data_only=True, read_only=True)
    parts = frontmatter(src)
    for ws in wb.worksheets:
        rows = []
        for row in ws.iter_rows(values_only=True):
            values = [clean_cell(cell) for cell in row]
            if any(values):
                rows.append(values)
        parts.extend([f"## Sheet｜{ws.title}", "", markdown_table(rows), "", "------------------------------------------------------------------------", ""])
    return "\n".join(parts).rstrip() + "\n"


def convert_delimited(src: Path, delimiter: str) -> str:
    text = src.read_text(encoding="utf-8-sig", errors="replace")
    reader = csv.reader(text.splitlines(), delimiter=delimiter)
    rows = [[clean_cell(cell) for cell in row] for row in reader]
    parts = frontmatter(src)
    parts.extend(["## 二、表格", "", markdown_table(rows), ""])
    return "\n".join(parts).rstrip() + "\n"


def convert_txt(src: Path) -> str:
    text = src.read_text(encoding="utf-8", errors="replace")
    parts = frontmatter(src)
    parts.extend(["## 二、正文", "", text.strip(), ""])
    return "\n".join(parts).rstrip() + "\n"


def convert_rtf(src: Path) -> str:
    raw = src.read_text(encoding="utf-8", errors="replace")
    raw = re.sub(r"\\'[0-9a-fA-F]{2}", "", raw)
    raw = re.sub(r"\\[a-zA-Z]+-?\d* ?", "", raw)
    raw = raw.replace("{", "").replace("}", "")
    raw = re.sub(r"\n{3,}", "\n\n", raw).strip()
    parts = frontmatter(src)
    parts.extend(["## 二、正文", "", raw, ""])
    return "\n".join(parts).rstrip() + "\n"


def convert_pdf(src: Path) -> str:
    reader = PdfReader(str(src))
    parts = frontmatter(src)
    for index, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        parts.extend([f"## 第 {index} 页", "", text.strip() or "_本页未提取到文字_", ""])
    return "\n".join(parts).rstrip() + "\n"


def convert_file(src: Path) -> tuple[bool, str]:
    suffix = src.suffix.lower()
    if src.name.startswith(".") or src.name == Path(__file__).name or suffix == ".md":
        return False, ""
    if suffix not in SUPPORTED:
        return False, f"不支持格式：{src.name}"
    if src.parent.name in SKIP_DIRS:
        return False, ""

    if suffix == ".docx":
        markdown = convert_docx(src)
    elif suffix in {".xlsx", ".xlsm"}:
        markdown = convert_xlsx(src)
    elif suffix == ".csv":
        markdown = convert_delimited(src, ",")
    elif suffix == ".tsv":
        markdown = convert_delimited(src, "\t")
    elif suffix == ".txt":
        markdown = convert_txt(src)
    elif suffix == ".rtf":
        markdown = convert_rtf(src)
    elif suffix == ".pdf":
        markdown = convert_pdf(src)
    else:
        return False, f"不支持格式：{src.name}"

    OUT_DIR.mkdir(exist_ok=True)
    dst = OUT_DIR / f"{src.stem}.md"
    dst.write_text(markdown, encoding="utf-8")
    return True, f"{src.name} → 已转换/{dst.name}"


def main() -> int:
    converted = []
    skipped = []
    for src in sorted(ROOT.iterdir()):
        if src.is_dir():
            continue
        ok, message = convert_file(src)
        if ok:
            converted.append(message)
        else:
            skipped.append(message)

    print("转换完成")
    print("")
    print("已转换：")
    if converted:
        for item in converted:
            print(f"- {item}")
    else:
        print("- 无")
    print("")
    print("未处理：")
    visible_skipped = [item for item in skipped if item]
    if visible_skipped:
        for item in visible_skipped:
            print(f"- {item}")
    else:
        print("- 无")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
